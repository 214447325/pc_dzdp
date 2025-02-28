
import time
import random
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from pyquery import PyQuery as pq
from openpyxl import load_workbook, Workbook

option = Options()

option.debugger_address = "localhost:9222"

# chrome --remote-debugging-port=9222
# chrome.exe --remote-debugging-port=9222 --user-data-dir="C:\selenum\AutomationProfile"

# 打开地址
def openURL(page=1,num=2):
    driver = webdriver.Chrome(options=option)
    url = f"https://www.dianping.com/shanghai/ch10/p{page}"
    driver.get(url)
    time.sleep(3)
    content = driver.find_element(By.XPATH,'//*[@id="shop-all-list"]/ul')
    element_text = str(content.get_attribute('innerHTML'))
    time.sleep(5)
    # print(element_text)
    doc = pq(element_text)
    # wb = Workbook()  # 创建一个新的工作簿
    # ws = wb.active  # 获取默认的工作表
    # ws.title = "MySheet"  # 修改工作表名称
    # ws['A1'] = "Hello"  # 在A1单元格写入数据
    # wb.save("example.xlsx")  # 保存文件
    workbook = load_workbook('dzdp.xlsx')
    sheet = workbook.active
    try:
        for i in doc('li').items():
            try:
                id = i('.pic').find('a').attr['data-shopid']
                tit = i('.tit').text()
                review_num = i('.review-num').text()
                mean_price = i('.mean-price').text()
                cx = i('.tag-addr').children('a').eq(0).text()
                address = i('.tag-addr').children('a').eq(1).text()
                recommend = i('.recommend').children('a').text()
                pic = i('.pic').find('img').attr['data-src']
                print('---------')
                print(id)
                print(tit)
                print(review_num)
                print(mean_price)
                print(cx)
                print(address)
                print(recommend)
                print(pic)
                sheet['A' + str(num)] = id
                sheet['B' + str(num)] = '上海'
                sheet['C' + str(num)] = tit
                sheet['D' + str(num)] = review_num
                sheet['E' + str(num)] = mean_price
                sheet['F' + str(num)] = cx
                sheet['G' + str(num)] = address
                sheet['H' + str(num)] = recommend
                sheet['I' + str(num)] = pic
                sheet['I' + str(num)] = 'web'
                num = num + 1
            except Exception as e:
                print(e)
        workbook.save('dzdp.xlsx')
        time.sleep(20)
    except Exception as e:
        print(e)
    
    print('===============================')
    print(f'第{page}页，获取成功')
    print(num)
    page = page + 1
    if page == 51:
        return False
    else:
        # 生成5-60的随机数用作停留时间
        r_timer = random.randint(5, 60)
        print(f'停留时间为{r_timer}秒')
        time.sleep(r_timer)
        return openURL(page,num)
    # print(doc.text())


if __name__ == '__main__':
    print('开始获取数据')
    openURL(20,287)
