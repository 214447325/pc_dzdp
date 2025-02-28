
import json

import uuid
from mitmproxy import http
from mitmproxy.http import Request
from mitmproxy.http import Response
from openpyxl import load_workbook, Workbook

num = 2

def request(flow: http.HTTPFlow) -> None:
    pass

def response(flow: http.HTTPFlow) -> None:
    global num
    if flow.request.url.startswith('https://m.dianping.com/wxmapi/search'):
        print('------------')
        # print('请求返回的地址=======>',flow.request.url)
        # print("请求content->",flow.response.content)
        # 获取index
        try:
            url = flow.request.url
            # print(url)
            url1 = url.split('?')[1]
            # print(url1)
            s = url1.split('&')
            index = 0
            index_num = 0
            for i in s:
                # print(i)
                if i.startswith('start'):
                    index_num = i.split('=')[1]
                    index = int(index_num) 
                    index = index + 2
            # print(index)
            content = flow.response.content
            json_data = json.loads(content)

            # print(json_data)
            if json_data['code'] == 200:
                data = json_data['data']
                data_list = data['list']
                workbook = load_workbook('dzdpwx.xlsx')
                sheet = workbook.active
                # print(data_list)
                try:
                    for item in data_list:
                        try:
                            # print(item)
                            time_based_uuid = uuid.uuid1()
                            print(time_based_uuid)
                            shopInfo = item['shopInfo']
                            name = shopInfo['name']
                            regionName = ''
                            starScore = ''
                            reviewCount = ''
                            priceText = ''
                            categoryName = ''
                            recommendReason = ''
                            tagList = []
                            tagName = ''
                            shopDealInfos = []
                            shopDealInfosName = ''
                            defaultPic = ''
                            
                            try:
                                if shopInfo['branchName'] != None and len(shopInfo['branchName']) > 0:
                                    print('1111111')
                                    name = name + shopInfo['branchName']
                            except Exception as e:
                                print(e)

                            try:
                                if shopInfo['regionName'] != None and len(shopInfo['regionName']) > 0:
                                    print('22222222')
                                    regionName = shopInfo['regionName']
                            except Exception as e:
                                print(e)

                            try:
                                if shopInfo['starScore'] != None and len(shopInfo['starScore']) > 0:
                                    print('333333')
                                    starScore = shopInfo['starScore']
                            except Exception as e:
                                print(e)

                            try:
                                if shopInfo['reviewCount'] != None and len(shopInfo['reviewCount']) > 0:
                                    print('44444444')
                                    reviewCount = shopInfo['reviewCount']
                            except Exception as e:
                                print(e)    

                            try:
                                if shopInfo['priceText'] != None and len(shopInfo['priceText']) > 0:
                                    print('5555555555')
                                    priceText = shopInfo['priceText']
                            except Exception as e:
                                print(e)
                            
                            try:
                                if shopInfo['categoryName'] != None and len(shopInfo['categoryName']) > 0:
                                    print('666666666')
                                    categoryName = shopInfo['categoryName']
                            except Exception as e:
                                print(e)



                            try:
                                if shopInfo['recommendReason'] != None and shopInfo['recommendReason'] != {}:
                                    if shopInfo['recommendReason']['text'] != None and len(shopInfo['recommendReason']['text']) > 0:
                                        print('777777')
                                        recommendReason = shopInfo['recommendReason']['text']
                            except Exception as e:
                                print(e)



                            try:
                               if shopInfo['tagList'] != None and len(shopInfo['tagList']) > 0:
                                    for k in shopInfo['tagList']:
                                        if k['text'] != None and len(k['text']) > 0:
                                            tagList.append(k['text'])
                                    if len(tagList) > 0:
                                        tagName = '、'.join(tagList)
                                    
                            except Exception as e:
                                print(e)

                            
                            try:
                                if shopInfo['shopDealInfos'] != None and len(shopInfo['shopDealInfos']) > 0:
                                    print('9999999999')
                                    if shopInfo['shopDealInfos'] != None and len(shopInfo['shopDealInfos']) > 0:
                                        for k in shopInfo['shopDealInfos']:
                                        # print('===========================')
                                        # print(k)
                                            if k['dealTitle'] != None:
                                                shopDealInfos.append(k['dealTitle'])
                                        if len(shopDealInfos) > 0:
                                            shopDealInfosName = '、'.join(shopDealInfos)

                                    
                            except  Exception as e:
                                print(e)

                            
                            if shopInfo['defaultPic'] != None and len(shopInfo['defaultPic']) > 0:
                                defaultPic = shopInfo['defaultPic']

                            # sheet['A' + str(index)] = time_based_uuid
                            sheet['A' + str(index)] = name
                            sheet['B' + str(index)] = '上海'
                            sheet['C' + str(index)] = regionName
                            sheet['D' + str(index)] = starScore
                            sheet['E' + str(index)] = reviewCount
                            sheet['F' + str(index)] = priceText
                            sheet['G' + str(index)] = categoryName
                            sheet['H' + str(index)] = recommendReason
                            sheet['I' + str(index)] = tagName
                            sheet['J' + str(index)] = shopDealInfosName
                            sheet['K' + str(index)] = defaultPic
                            sheet['L' + str(index)] = 'wx'
                            sheet['M' + str(index)] = index_num
                            sheet['N' + str(index)] = "静安区"
                            print(f"名称:{name} =>门店:{regionName} =>评分:{starScore} =>评论数:{reviewCount} =>人均消费:{priceText} =>分类:{categoryName} =>描述:{recommendReason} =>榜单:{tagName} =>优惠:{shopDealInfosName} =>图片:{defaultPic}")
                            # num = int(num) + 1
                        except Exception as e:
                            print(e)
                            pass
                        index = index + 1
                except Exception as e:
                    print(e)

                workbook.save('dzdpwx.xlsx')
                print('===========================保存成功=========================')
        except Exception as e:
            print(e)
            pass


        # content1 = flow.request.content.decode('utf-8')
        # print('==============')
        # print(content1)
        # s = content1.split('&')
        # index = 0
        # for i in s:
        #     if i.startswith('start'):
        #         index = i.split('=')[1]     
        # # 获取内容
        # try:
        #     content = flow.response.content
        #     json_data = json.loads(content)
        #     print(json_data['code'])
        #     if json_data['code'] == 0:
        #         time_based_uuid = uuid.uuid1()
        #         print(time_based_uuid)
        #         data = json_data['data']
        #         poilist = data['poilist']
        #         for item in poilist:
        #             print(item)
        # except Exception as e:
        #     print(e)
        # print(index)

        # workbook = load_workbook('mt.xlsx')
        # sheet = workbook.active
        # # sheet['A' + str(index)] = index
        # # sheet['B' + str(index)] = '上海'
        # # sheet['C' + str(index)] = json_data
        # workbook.save('mt.xlsx')
        # print('数据保存成功！')
       
    